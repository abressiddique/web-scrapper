from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import sys

def scrape_phone_options(phone_name):
    chrome_options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=chrome_options)

    try:
        # Open the MySmartPrice website
        driver.get("https://www.mysmartprice.com/mobile/")
        
        # Find the search bar by name
        search_bar = driver.find_element(By.NAME, "s")
        
        # Type the phone name into the search bar
        search_bar.send_keys(phone_name)
        
        # Simulate pressing the Enter key to initiate the search
        search_bar.send_keys(Keys.RETURN)
        
        # Wait for the page to load
        driver.implicitly_wait(10)
        
        # Find the first 8 items and retrieve their names and URLs using XPath
        phone_options = driver.find_elements(By.XPATH, "//a[@class='prdct-item__name']")
        
        if not phone_options:
            print("No phone options found. Check if the website structure has changed.")
            driver.quit()
            return [], []
        
        phone_names = [option.text.strip() for option in phone_options[:8]]
        phone_urls = [option.get_attribute("href") for option in phone_options[:8]]
        
    except Exception as e:
        print(f"An error occurred: {e}")
        phone_names, phone_urls = [], []
        
    finally:
        driver.quit()

    return phone_names, phone_urls


def choose_phone_to_scrape(phone_names):
    print("Choose a phone to scrape:")
    for i, name in enumerate(phone_names, start=1):
        print(f"{i}. {name}")

    choice = int(input("Enter the number corresponding to your choice: "))
    return phone_names[choice - 1] if 1 <= choice <= len(phone_names) else None


def scrape_phone_details(phone_url):
    try:
        # Send an HTTP GET request to the phone's details page
        r = requests.get(phone_url)
        
        # Check if the request was successful
        if r.status_code == 200:
            # Parse the HTML content of the phone's details page
            soup = BeautifulSoup(r.text, "html.parser")
            
            # Extract the specific information you mentioned
            phone_name = soup.find("h1", class_="prdct-dtl__ttl").text.strip()
            price = soup.find("div", class_="prdct-dtl__prc").find("span", class_="prdct-dtl__prc-val").text.strip()
            
            key_specifications = []
            key_spec_list = soup.find("ul", class_="kyspc__list clearfix")
            key_specs = key_spec_list.find_all("li", class_="kyspc__item")
            for key_spec in key_specs:
                spec_text = key_spec.find("span").text.strip()
                spec_title = key_spec.text.strip().replace(spec_text, "").strip()
                key_specifications.append((spec_text, spec_title))
            
            amazon_link = soup.find("span", class_="js-open-link")["data-open-link"]
            
            return phone_name, price, key_specifications, amazon_link
        else:
            print(f"Failed to retrieve phone details. Status code: {r.status_code}")
            return None, None, None, None
            
    except Exception as e:
        print(f"An error occurred while scraping phone details: {e}")
        return None, None, None, None


def insert_data_into_excel(extracted_data, excel_file_path):
    try:
        # Load the existing workbook
        wb = load_workbook(excel_file_path)
        ws = wb.active
    except FileNotFoundError:
        # If the workbook doesn't exist, create a new one
        wb = Workbook()
        ws = wb.active

    if extracted_data:
        phone_name, price, key_specifications, amazon_link = extracted_data

        if phone_name:
            ws.append(["Phone Name", phone_name])
            ws.append(["Price", price])

            for title, text in key_specifications:
                ws.append([title, text])

            ws.append(["Amazon Link", amazon_link])

        # Save the workbook
        wb.save(excel_file_path)

    # Close the workbook
    wb.close()

if __name__ == "__main__":
    # Check if the correct number of command-line arguments is provided
    if len(sys.argv) != 2:
        print("Usage: python Scrapper(Retrieval).py <inputValue>")
        sys.exit(1)

    # Retrieve the input value from the command line
    input_value = sys.argv[1]

    phone_name = input_value

    # Scrape phone options and let the user choose
    phone_names, phone_urls = scrape_phone_options(phone_name)
    
    if not phone_names:
        print("No phone names found. Exiting.")
        sys.exit(1)
    
    selected_phone_name = choose_phone_to_scrape(phone_names)

    if selected_phone_name:
        # Get the corresponding URL for the selected phone
        selected_phone_url = phone_urls[phone_names.index(selected_phone_name)]

        # Scrape detailed information for the selected phone
        data = scrape_phone_details(selected_phone_url)
        
        if not data[0]:
            print("Failed to scrape phone details. Exiting.")
            sys.exit(1)
        
        # Save the data to the Excel workbook
        excel_file_path = "C:\\Users\\91986\\Desktop\\web scrapper\\Book.xlsx"
        insert_data_into_excel(data, excel_file_path)
        print(f"Scraped data saved to {excel_file_path}")
    else:
        print("Invalid choice. Exiting.")
